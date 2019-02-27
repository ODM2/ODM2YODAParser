import os
import re
from collections import defaultdict
from uuid import uuid4
import wx
from datetime import datetime

from pubsub import pub
from pandas import isnull, DataFrame, NaT
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm.exc import NoResultFound
from sqlalchemy.orm.session import Session
import openpyxl
from openpyxl.worksheet.table import Table
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell
from yodatools.excelparser.ParserException import ParserException

from odm2api.models import \
    (Base,
     DataSets,
     Citations,
     AuthorLists,
     People,
     Units,
     SamplingFeatures,
     Organizations,
     Affiliations,
     ProcessingLevels,
     Sites,
     SpatialReferences,
     Methods,
     Variables,
     Actions,
     FeatureActions,
     ActionBy,
     TimeSeriesResults,
     DataSetsResults,
     TimeSeriesResultValues,
     CVUnitsType,
     CVVariableName,
     setSchema)


class ExcelParser(object):

    TABLE_NAMES = [
        'Analysis_Results',
        'DatasetCitation'
        'Organizations',
        'People'
        'ProcessingLevels',
        'Sites',
        'SpatialReferences',
        'SpecimenAnalysisMethods',
        'SpecimenCollectionMethods',
        'Specimens',
        'Units',
        'Varibles'
    ]

    def __init__(self, input_file, session_factory, **kwargs):

        self.input_file = input_file

        self.__session_factory = session_factory

        self.session = session_factory.getSession()  # type: Session
        self.engine = session_factory.engine

        self.total_rows_to_read = 0
        self.rows_read = 0
        self.workbook = None
        self.sheets = []
        self.name_ranges = {}
        self.tables = {}

        self.orgs = defaultdict(lambda: None)
        self.affiliations = defaultdict(lambda: None)
        self.data_set = defaultdict(lambda: None)
        self.methods = defaultdict(lambda: None)
        self.variables = defaultdict(lambda: None)
        self.units = defaultdict(lambda: None)
        self.processing_levels = defaultdict(lambda: None)
        self.spatial_references = defaultdict(lambda: None)

        self._init_data(input_file)

    def _init_data(self, file_path):

        self.update_progress_label('Loading %s' % file_path)

        self.workbook = openpyxl.load_workbook(file_path, data_only=True)  # type: Workbook

        # Loop through worksheets to grab table data
        for ws in self.workbook.worksheets:

            tables = getattr(ws, '_tables', [])
            for table in tables:  # type: Table

                if table.name in ['AuthorList', 'ExternalIDOrgs', 'ControlledVocabularies', 'ExternalIdentifiers']:
                    # skip these tables because they do not (currently) need to be parsed
                    # and they mess up the total row count calculation
                    continue

                self.update_progress_label('Loading table data: %s' % table.name)

                rows = ws[table.ref]

                # check if table_rows length is less than 2, since the first row is just the table headers
                # if True, then the current table has no data
                if len(rows) < 2:
                    continue

                # get headers from row 1
                headers = map(lambda x: x.replace('[CV]', '').strip(), [cell.value for cell in rows[0]])

                # get values from rows 2...n
                data = [[cell.value for cell in row] for row in rows[1:]]

                self.tables[table.name.strip()] = DataFrame(data, columns=headers).dropna(how='all')

        self.update_progress_label('Calculating total row size')
        self.total_rows_to_read = sum([table.shape[0] for table in self.tables.values()])

    def get_or_create(self, model, values, check_fields=None, filter_by=None, commit=True):  # type: (Base, dict, [str], str|[str], bool) -> Base
        """
        Gets an existing instance of <model> or creates a new one if not found

        :param model: The model from odm2api.models used to create the object
        :param values: A dict containing the fields to insert into the database (given the record does not exist).
        :param check_fields: A list of strings of required field names (optional).
        :param filter_by: A string or list of strings used to filter queries by. If None, the query will filter using **values (optional).
        :param commit: Boolean value indicating whether or not to commit the transaction.
        :return: An instance of the retrieved or created model.
        :raise ValueError: Raised when a value in values is NaT given the key exists in check_fields
        """
        if check_fields:
            bad_fields = []
            for field in check_fields:
                if isnull(values[field]):
                    bad_fields.append(field)

            if len(bad_fields):
                raise ValueError('Object "{}" is missing required fields: {}'.format(model.__tablename__.title(),
                                                                                     ', '.join(bad_fields)))

        filters = {}

        if isinstance(filter_by, str):
            filters[filter_by] = values.get(filter_by, None)
        elif isinstance(filter_by, list):
            for f in filter_by:
                filters[f] = values.get(f, None)
        else:
            filters = values

        instance = self.get(model, **filters)

        if instance:
            return instance
        else:
            return self.create(model, commit=commit, **values)

    def create(self, model, commit=True, **kwargs):
        """
        Creates an instance of <model>

        :param model: an ODM2 model
        :param commit: boolean, commits the newly created object if true
        :param kwargs: keyword arguments used to create <model>
        :return:
        """
        instance = model(**kwargs)
        self.session.add(instance)

        if commit:
            self.session.commit()

        return instance

    def get(self, model, **kwargs):
        """
        Gets a single instance of an ODM2 model

        :param model: class of the model to query
        :param kwargs: values to use in query
        :return: an instance of <model>
        """
        try:
            return self.session.query(model).filter_by(**kwargs).one()
        except NoResultFound:
            return None


    def _flush(self):
        try:
            self.session.flush()
        except IntegrityError as e:

            if os.getenv('DEBUG') == 'true':
                print(e)

            self.session.rollback()

    def update_progress_label(self, message, label_pos=1):
        pub.sendMessage('controller.update_progress_label', message=message, label_pos=label_pos)

    def update_output_text(self, message):
        pub.sendMessage('controller.update_output_text', message='%s\n' % message)

    def update_gauge(self, rows_read=1, message=None, gauge_pos=1, label_pos=1, setvalue=None):
        """
        Updates the gauge based on `self.rows_read`
        :return: None
        """
        if message is not None:
            self.update_progress_label(message, label_pos=label_pos)

        if setvalue is not None:
            value = setvalue

        else:

            self.rows_read += rows_read

            try:
                value = (float(self.rows_read) / float(self.total_rows_to_read)) * 100.0
            except ZeroDivisionError:
                return

        pub.sendMessage('controller.update_gauge', value=value, gauge_pos=gauge_pos)

    def get_named_range(self, sheet, coord):
        """
        Returns the range of cells contained in a given worksheet by a given set of coordinates
        :param sheet: string like.
            Name of the worksheet
        :param coord: string like
            String representation of sheet coordinates
        :return: Range of cell(s)
        """
        ws = self.workbook[sheet]
        return ws[coord]

    def get_named_range_value(self, sheet, coord):
        """
        Gets the value of the cell(s) in a given worksheet at a given set of coordinates

        :param sheet: string like.
            Name of the worksheet with the named range
        :param coord: string like.
            String representation of the named range coordinate (e.g. '$A$1')
        :return: Value(s) contained in the named range given by `coord`
        """

        value = self.get_named_range(sheet, coord)

        if isinstance(value, tuple):
            results = []
            for v in value:
                results.append(v[0].value)
            # value = [v.value for v in value]
            return results
        elif hasattr(value, 'value'):
            return value.value

        return value

    def get_named_range_cell_value(self, named_range):
        """
        Gets the value of the cell given by named_range. The passed in named range
        should reference only a single cell.


        :param named_range: string like.
            Name of the named range
        :return:
        """
        try:
            nr = self.workbook.defined_names[named_range]
            return self.get_named_range_value(*next(nr.destinations))
        except KeyError:
            return None


    def parse_name(self, fullname):  # type: (str) -> dict
        """
        Parses a full name contained in a string and returns a dict representation
        of the name. Also removes trailing/leading whitespace of the names.

        If `fullname` does not contain a comma, it's assumed `fullname` is formatted as:

            "<first name> <middles name(s)> <last name>"

        If `fullname` contains a comma (e.g. "Doe, John"), then it is assumed `fullname`
        if formatted as:

            "<last name>, <first name> <middle name(s)>"

        :param fullname:
        :return:
        """
        values = re.split(r'\s+', fullname)

        if any([',' in name for name in values]):
            # `fullname` contained a comma (formatted as "<last>, <first> <middle>")
            # so do a little rearranging.
            lastname = values.pop(0).replace(',', '')

            try:
                firstname = values.pop(0)
            except IndexError:
                firstname = ''

            values = [firstname] + values + [lastname]

        names = {
            'first': values[0],
            'last': values[-1],
            'middle': ' '.join(values[1:-1])
        }

        return names

    def create_action(self, start_date, end_date, utcoffset, method, commit=False):  # type: (datetime, datetime, int, Methods, bool) -> Actions
        """
        Creates an ODM2 Actions object

        :param start_date: datetime like
        :param end_date: datetime like
        :param utcoffset: int like
        :param method: Methods object
        :param commit: bool
        :return:
        """

        utcoffset = int(utcoffset)

        return self.create(Actions, commit=commit, **{
            'MethodObj': method,
            'ActionTypeCV': "Observation",
            'BeginDateTime': start_date,
            'BeginDateTimeUTCOffset': utcoffset,
            'EndDateTime': end_date,
            'EndDateTimeUTCOffset': -7
        })

    def create_feature_action(self, sampling_feature, action, commit=False):  # type: (SamplingFeatures, Actions, bool) -> FeatureActions
        return self.create(FeatureActions, commit=commit, **{
            'SamplingFeatureObj': sampling_feature,
            'ActionObj': action
        })

    def create_action_by(self, affiliation, action, commit=False):  # type: (Affiliations, Actions, bool) -> ActionBy
        return self.create(ActionBy, commit=commit, **{
            'AffiliationObj': affiliation,
            'ActionObj': action,
            'IsActionLead': True
        })

    def parse_people_and_orgs(self):

        self.update_progress_label('Reading Organizations')

        organization_table = self.tables.get('Organizations', DataFrame())
        for _, row in organization_table.iterrows():
            params = {
                'OrganizationTypeCV': row.get('Organization Type'),
                'OrganizationCode': row.get('Organization Code'),
                'OrganizationName': row.get('Organization Name')
            }

            # check if params has required fields
            assert all(params.values()), 'Values = %s ' % str(params.values())

            # add non required fields
            params.update(OrganizationLink=row.get('Organization Link', None),
                          OrganizationDescription=row.get('Organization Description', None))

            org = self.get_or_create(Organizations, params, filter_by='OrganizationName', commit=False)
            self.orgs[row.get('Organization Name')] = org  # save this for later when we create Affiliations

            self.update_gauge()

        self.session.commit()


        # Create Person and Affiliation objects
        self.update_progress_label('Reading People')

        people_table = self.tables.get('People', DataFrame())
        for _, row in people_table.iterrows():  # type: (any, DataFrame)

            row.fillna(value='', inplace=True)  # replace NaN values with empty string

            person_params = {
                'PersonFirstName': row.get('First Name'),
                'PersonLastName': row.get('Last Name'),
                'PersonMiddleName': row.get('Middle Name')
            }

            if NaT in person_params.values():
                continue

            person = self.get_or_create(People, person_params)

            aff_params = {
                'AffiliationStartDate': row.get('Affiliation Start Date'),
                'AffiliationEndDate': row.get('Affiliation End Date'),
                'PrimaryPhone': row.get('Primary Phone'),
                'PrimaryEmail': row.get('Primary Email'),
                'PrimaryAddress': row.get('Primary Address'),
                'PersonLink': row.get('Person Link'),
                'OrganizationObj': self.orgs.get(row.get('Organization Name')),
                'PersonObj': person
            }

            start_date = aff_params['AffiliationStartDate']
            aff_params['AffiliationStartDate'] = datetime(year=start_date.year, month=start_date.month,
                                                          day=start_date.day, hour=start_date.hour,
                                                          minute=start_date.minute, second=start_date.second)

            del aff_params['AffiliationEndDate']

            aff = self.get_or_create(Affiliations, aff_params, filter_by='PersonID')
            self.affiliations[row.get('Full Name')] = aff

            self.update_gauge()

    def parse_datasets(self):

        self.update_progress_label('parsing datasets')

        dataset_uuid = self.get_named_range_cell_value('DatasetUUID')
        dataset_type = self.get_named_range_cell_value('DatasetType')
        dataset_code = self.get_named_range_cell_value('DatasetCode')
        dataset_title = self.get_named_range_cell_value('DatasetTitle')
        dataset_abstract = self.get_named_range_cell_value('DatasetAbstract')

        params = {
            'DataSetUUID': dataset_uuid,
            'DataSetTypeCV': dataset_type,
            'DataSetCode': dataset_code,
            'DataSetTitle': dataset_title,
            'DataSetAbstract': dataset_abstract
        }

        self.data_set = self.get_or_create(DataSets, params, filter_by=['DataSetCode'])

    def parse_methods(self, table=None):
        """
        Parse Methods recorded in the excel template

        :param table: A dataframe containing the Method table data
        :return: None
        """

        if table is None:
            table = self.tables.get('Methods', DataFrame())  # type: DataFrame

        # Force values in 'Method Code' column to be strings
        table['Method Code'] = table['Method Code'].astype(str)

        self.update_progress_label('Reading Methods table')

        for _, row in table.iterrows():

            self.methods[row.get('Method Code', '').lower()] = self.parse_method(**row)

        self.session.commit()

        self.update_gauge(table.shape[0])

    def parse_method(self, **kwargs):

        org = self.orgs.get(kwargs.get('Organization Name'))

        params = {
            'MethodTypeCV': kwargs.get('Method Type'),
            'MethodCode': kwargs.get('Method Code'),
            'MethodName': kwargs.get('Method Name')
        }

        # check if params has required fields
        assert all(params.values()), 'Values = %s ' % str(params.values())

        # After checking for required fields, add the non required field
        params.update(MethodLink=kwargs.get('MethodLink'),
                      MethodDescription=kwargs.get('Method Description'),
                      OrganizationObj=org)

        return self.get_or_create(Methods, params, filter_by='MethodCode', commit=False)

    def parse_variables(self):
        table = self.tables.get('Variables', DataFrame())

        table.replace({'NULL': None}, inplace=True)

        self.update_progress_label('Reading Variables table')

        for _, row in table.iterrows():

            params = {
                'VariableTypeCV': row.get('Variable Type'),
                'VariableCode': row.get('Variable Code'),
                'VariableNameCV': row.get('Variable Name'),
                'NoDataValue': row.get('No Data Value')
            }

            assert(all(params.values()))

            params.update(VariableDefinition=row.get('Variable Definition'),
                          SpeciationCV=row.get('Speciation'))

            variable = self.get_or_create(Variables, params, filter_by=['VariableCode'], check_fields=['NoDataValue'], commit=False)
            self.variables[params.get('VariableCode').lower()] = variable

        self.session.commit()

        self.update_gauge(table.shape[0])

    def parse_units(self):
        self.update_progress_label('Reading Units')

        table = self.tables.get('Units', DataFrame())
        for _, row in table.iterrows():

            params = {
                'UnitsTypeCV': row.get('Units Type'),
                'UnitsAbbreviation': row.get('Units Abbreviation'),
                'UnitsName': row.get('Units Name')
            }

            assert(all(params.values()))

            params.update(UnitsLink=row.get('Units Link'))

            unit = self.get_or_create(Units, params, filter_by=['UnitsName', 'UnitsAbbreviation', 'UnitsTypeCV'],
                                      check_fields=['UnitsTypeCV'])
            self.units[params.get('UnitsName').lower()] = unit

        self.update_gauge(table.shape[0])

    def parse_spatial_reference(self):
        """
        Parse spatial references
        :return: None
        """

        self.update_progress_label('Reading SpatialReferences table')

        table = self.tables.get('SpatialReferences', DataFrame())
        for _, row in table.iterrows():

            params = {
                'SRSCode': row.get('SRSCode'),
                'SRSName': row.get('SRSName'),
                'SRSDescription': row.get('SRSDescription'),
                'SRSLink': row.get('SRSLink'),
            }

            assert(params.get('SRSName'))

            sref = self.get_or_create(SpatialReferences, params, filter_by=['SRSCode'], commit=False)
            self.spatial_references[row.get('SRSName', '').lower()] = sref

        self.session.commit()

    def parse_processing_level(self):

        self.update_progress_label('Reading ProcessingLevels table')

        # processing_codes = self.get_named_range_cell_value('ProcessingLevelCodes')
        # processing_codes = [code for code in processing_codes if code is not None]
        table = self.tables.get('ProcessingLevels', DataFrame())

        if 'Processing Level Code' not in table.keys():
            raise ParserException('Processing Level Codes not found. (Processing Level Information probably not formatted as a table in excel)')


        table['ProcessingLevelCodes'] = table['Processing Level Code'].astype(int).astype(str)


        for _, row in table.iterrows():

            params = {
                'ProcessingLevelCode': str(int(row.get('Processing Level Code'))),
                'Definition': row.get('Definition'),
                'Explanation': row.get('Explanation')
            }

            # assert(params.get('ProcessingLevelCode', False))

            plvl = self.get_or_create(ProcessingLevels, params, filter_by=['ProcessingLevelCode'])
            self.processing_levels[params.get('ProcessingLevelCode')] = plvl

            self.update_gauge()

        self.session.commit()

    def get_table_name_ranges(self):
        """
        Returns a list of the name range that have a table.
        The name range should contain the cells locations of the data.
        :rtype: list
        """
        CONST_NAME = "_Table"
        table_name_range = {}
        for name_range in self.name_ranges:
            if CONST_NAME in name_range.name:
                sheet = name_range.attr_text.split('!')[0]
                sheet = sheet.replace('\'', '')

                if sheet in table_name_range:
                    table_name_range[sheet].append(name_range)
                else:
                    table_name_range[sheet] = [name_range]

        return table_name_range

    def get_range_address(self, named_range):
        """
        Depracated

        :param named_range:
        :return:
        """
        if named_range is not None:
            return named_range.attr_text.split('!')[1].replace('$', '')
        return None

    def get_range_value(self, range_name, sheet):

        """
        Depracated

        :param range_name:
        :param sheet:
        :return:
        """
        value = None
        named_range = self.workbook.get_named_range(range_name)
        range_ = self.get_range_address(named_range)
        if range_:
            value = sheet[range_].value
        return value

    def get_sheet_and_table(self, sheet_name):
        """
        Depracated

        :param sheet_name:
        :return:
        """
        if sheet_name not in self.tables:
            return [], []
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        tables = self.tables[sheet_name]

        return sheet, tables